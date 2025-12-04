# app.py (LIMPIO Y FUNCIONAL)

from flask import (
    Flask,
    render_template,
    request,
    redirect,
    url_for,
    flash,
    Response,
    send_from_directory,
    abort,
    jsonify,
    make_response,
    current_app,
)
from flask_login import (
    LoginManager,
    login_user,
    logout_user,
    current_user,
    login_required,
)

from datetime import datetime, date, timedelta
import zoneinfo

# ==============================
# ZONA HORARIA (SANTO DOMINGO)
# ==============================
LOCAL_TZ = zoneinfo.ZoneInfo("America/Santo_Domingo")


def now_local() -> datetime:
    """Fecha y hora actual en Santo Domingo."""
    return datetime.now(LOCAL_TZ)


def today_local() -> date:
    """Fecha actual (date) en Santo Domingo."""
    return now_local().date()


anio_actual = now_local().year
mes_actual = now_local().month

# Para generar la gráfica en imagen
import matplotlib

matplotlib.use("Agg")  # importante: backend sin interfaz gráfica
import matplotlib.pyplot as plt
import uuid

from io import StringIO, BytesIO
from functools import wraps
import csv, json, os
from dateutil import parser

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter

from sqlalchemy import text
from sqlalchemy.orm import joinedload

from xhtml2pdf import pisa
from werkzeug.utils import secure_filename

from models import db, User, Hospital, EmergencyRecord, GuardiaEmergencia, Internamiento


# ==============================
# APP CONFIG
# ==============================
app = Flask(__name__)


# ==============================
# LOGOS CONFIG
# ==============================
LOGOS_SUBFOLDER = "img"  # dentro de /static
UPLOAD_FOLDER_LOGOS = os.path.join(app.root_path, "static", LOGOS_SUBFOLDER)
os.makedirs(UPLOAD_FOLDER_LOGOS, exist_ok=True)

ALLOWED_LOGO_EXTENSIONS = {"png", "jpg", "jpeg", "gif"}


def allowed_logo_file(filename: str) -> bool:
    if not filename:
        return False
    return (
        "." in filename
        and filename.rsplit(".", 1)[-1].lower() in ALLOWED_LOGO_EXTENSIONS
    )


# ==============================
# DB CONFIG (MySQL / Railway)
# ==============================
DATABASE_URL = os.getenv(
    "DATABASE_URL", "mysql+pymysql://root:@127.0.0.1:3306/emergencias?charset=utf8mb4"
)

if not DATABASE_URL:
    rh = os.getenv("MYSQLHOST")
    ru = os.getenv("MYSQLUSER")
    rp = os.getenv("MYSQLPASSWORD")
    rport = os.getenv("MYSQLPORT", "3306")
    rdb = os.getenv("MYSQLDATABASE")
    if rh and ru and rp and rdb:
        DATABASE_URL = f"mysql+pymysql://{ru}:{rp}@{rh}:{rport}/{rdb}?charset=utf8mb4"

if not DATABASE_URL:
    raise RuntimeError(
        "No se ha configurado la base de datos. "
        "Define DATABASE_URL o las variables MYSQLHOST, MYSQLUSER, "
        "MYSQLPASSWORD, MYSQLPORT, MYSQLDATABASE."
    )

if DATABASE_URL.startswith("mysql://"):
    DATABASE_URL = DATABASE_URL.replace("mysql://", "mysql+pymysql://", 1)

if DATABASE_URL.startswith("mysql+pymysql://") and "charset=" not in DATABASE_URL:
    sep = "&" if "?" in DATABASE_URL else "?"
    DATABASE_URL = f"{DATABASE_URL}{sep}charset=utf8mb4"

app.config["SQLALCHEMY_DATABASE_URI"] = DATABASE_URL
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["SECRET_KEY"] = os.getenv("SECRET_KEY", "cambia-esta-clave")
app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {"pool_pre_ping": True}

db.init_app(app)
from bootstrap import run_bootstrap_on_start

run_bootstrap_on_start(app)

# ==============================
# LOGIN MANAGER
# ==============================
login_manager = LoginManager()
login_manager.login_view = "login"
login_manager.init_app(app)


@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))


# ==============================
# DECORADORES DE PERMISOS
# ==============================
def admin_required(fn):
    """Solo Admin general (is_admin=True)."""

    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin:
            flash("Solo Administradores generales.", "danger")
            return redirect(url_for("index"))
        return fn(*args, **kwargs)

    return wrapper


def hospital_admin_required(fn):
    """
    Permite acceso a:
      - Admin general (is_admin=True)
      - Admin de hospital (is_hospital_admin=True)
    """

    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for("login"))
        if not (current_user.is_admin or current_user.is_hospital_admin):
            flash("Solo administradores (general o de hospital).", "danger")
            return redirect(url_for("index"))
        return fn(*args, **kwargs)

    return wrapper


def user_hospital_scope():
    """
    Hospital filtro según rol:
    - Admin general: None (ve todos)
    - Admin hospital/usuario normal: su hospital
    """
    if current_user.is_admin:
        return None
    return current_user.hospital


# ==============================
# CONTEXT PROCESSOR
# ==============================
@app.context_processor
def inject_choices():
    """Hace disponibles los hospitales activos en todas las plantillas."""
    try:
        hospitales = (
            Hospital.query.filter_by(activo=True).order_by(Hospital.nombre.asc()).all()
        )
    except Exception:
        hospitales = []
    return dict(HOSPITALES=hospitales)


# ==============================
# BOOTSTRAP DB (OPCIONAL)
# ==============================
def _seed_hospitals():
    HOSPITALES_BASE = [
        "Hospital Regional Juan Pablo Pina",
        "Hospital Provincial Dr. Rafael j Mañón",
        "Hospital Provincial Nuestra señora de regla",
        "Hospital Municpal Villa Fundacion",
        "Hospital Municipal Barsequillo",
        "Hospital Municipal Maria Paniagua",
        "Hospital Municipal Tomasina Valdez",
        "Hospital Municipal Nizao",
        "Hospital  Municipal Cambita pueblo",
        "Hospital Municipal Cambita Garabitos",
        "Hospital Municipal de Yaguate",
        "Hospital Municipal Villa Altagracia",
        "Hospital Nustra Señora de Altagracia",
        "Hospital Municipal Dr.Guarionex ALcantara",
        "Hospital Provincial San José de Ocoa",
        "Hospital Municipal los Cacaos",
    ]
    creados = 0
    for nombre in HOSPITALES_BASE:
        if not Hospital.query.filter_by(nombre=nombre).first():
            db.session.add(Hospital(nombre=nombre, activo=True))
            creados += 1
    db.session.commit()
    print(f"[BOOTSTRAP] Hospitales OK (nuevos: {creados})")


def bootstrap_if_empty():
    """Crea tablas y un admin si la DB está vacía (BOOTSTRAP_ON_START=1)."""
    with app.app_context():
        db.create_all()
        total = User.query.count()
        print(f"[BOOTSTRAP] Usuarios existentes: {total}")
        if total == 0:
            _seed_hospitals()
            admin_user = os.getenv("ADMIN_USER", "admin")
            admin_pass = os.getenv("ADMIN_PASS", "Admin123*")
            admin_hosp = os.getenv("ADMIN_HOSPITAL", "Hospital Municipal los Cacaos")

            ok = Hospital.query.filter_by(nombre=admin_hosp, activo=True).first()
            if not ok:
                any_h = Hospital.query.filter_by(activo=True).first()
                admin_hosp = any_h.nombre if any_h else "Hospital Municipal los Cacaos"

            u = User(
                username=admin_user,
                hospital=admin_hosp,
                is_admin=True,
                is_hospital_admin=True,
            )
            u.set_password(admin_pass)
            db.session.add(u)
            db.session.commit()
            print(f"[BOOTSTRAP] Admin creado: {admin_user} / hospital={admin_hosp}")
        else:
            print("[BOOTSTRAP] Ya hay usuarios. No se crea admin nuevo.")


if os.getenv("BOOTSTRAP_ON_START", "0") == "1":
    try:
        bootstrap_if_empty()
    except Exception as e:
        print(f"[BOOTSTRAP] Error: {e}")


@app.route("/admin/bootstrap")
def admin_bootstrap():
    token = request.args.get("token", "")
    expected = os.getenv("SETUP_TOKEN", "")
    if not expected or token != expected:
        return abort(403)
    try:
        bootstrap_if_empty()
        return "Bootstrap ejecutado", 200
    except Exception as e:
        return f"Error: {e}", 500


# ==============================
# HEALTHCHECK
# ==============================
@app.route("/healthz")
def healthz():
    try:
        db.session.execute(text("SELECT 1"))
        return "OK", 200
    except Exception as e:
        return f"DB ERROR: {e}", 500


# ==============================
# PÁGINAS BASE / LOGIN / Perfil de Usuario
# ==============================
from flask import redirect, url_for
from flask_login import current_user

@app.route("/")
def index():
    # Si está autenticado, lo mandamos al dashboard
    if current_user.is_authenticated:
        return redirect(url_for("dashboard"))
    # Si no, al login (o a una landing si quieres)
    return redirect(url_for("login"))



@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        password = request.form.get("password") or ""
        user = User.query.filter_by(username=username).first()
        if user and user.check_password(password):
            login_user(user)
            flash("Sesión iniciada.", "success")
            return redirect(url_for("dashboard"))
        flash("Credenciales inválidas.", "danger")
    return render_template("login.html")


@app.route("/logout")
@login_required
def logout():
    logout_user()
    flash("Sesión cerrada.", "success")
    return redirect(url_for("index"))


@app.route("/perfil", methods=["GET", "POST"])
@login_required
def perfil():
    u = current_user

    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        password1 = request.form.get("password1") or ""
        password2 = request.form.get("password2") or ""

        if not username:
            flash("El nombre de usuario es obligatorio.", "danger")
            return render_template("perfil.html", u=u)

        existe = User.query.filter(User.id != u.id, User.username == username).first()
        if existe:
            flash("Ese nombre de usuario ya está en uso.", "danger")
            return render_template("perfil.html", u=u)

        u.username = username
        u.nombre = request.form.get("nombre")
        u.apellido = request.form.get("apellido")
        u.cedula = request.form.get("cedula")
        u.especialidad = request.form.get("especialidad")
        u.cargo = request.form.get("cargo")
        u.exequatur = request.form.get("exequatur")
        u.telefono = request.form.get("telefono")
        u.email = request.form.get("email")

        if password1 or password2:
            if password1 != password2:
                flash("Las contraseñas no coinciden.", "danger")
                return render_template("perfil.html", u=u)
            u.set_password(password1)

        db.session.commit()
        flash("Perfil actualizado correctamente.", "success")
        return redirect(url_for("perfil"))

    return render_template("perfil.html", u=u)


# ======================================================================
#   REGISTRO DIARIO DE EMERGENCIAS
# ======================================================================
@app.route("/nuevo", methods=["GET", "POST"])
@login_required
@hospital_admin_required
def nuevo():
    if request.method == "POST":
        try:
            fecha_str = request.form.get("fecha")
            fecha = parser.parse(fecha_str).date() if fecha_str else today_local()

            if current_user.is_admin:
                hospital_nombre = (request.form.get("hospital") or "").strip()
                ok = Hospital.query.filter_by(
                    nombre=hospital_nombre, activo=True
                ).first()
                if not ok:
                    flash("Hospital inválido o inactivo.", "danger")
                    return render_template("form.html")
                hospital = hospital_nombre
            else:
                hospital = current_user.hospital

            existente = EmergencyRecord.query.filter_by(
                fecha=fecha, hospital=hospital
            ).first()
            if existente:
                flash(
                    "Ya existe un registro para este hospital en esa fecha. "
                    "Edítalo en lugar de crear uno nuevo.",
                    "danger",
                )
                return redirect(url_for("editar", rec_id=existente.id))

            def to_int(name):
                val = (request.form.get(name) or "0").strip()
                try:
                    return max(int(val), 0)
                except Exception:
                    return 0

            rec = EmergencyRecord(
                fecha=fecha,
                hospital=hospital,
                atenciones=to_int("atenciones"),
                ingresos=to_int("ingresos"),
                alta_voluntario=to_int("alta_voluntario"),
                traslados=to_int("traslados"),
                defunciones=to_int("defunciones"),
                motivo_traslado=(request.form.get("motivo_traslado") or "").strip(),
                hospital_referencia=(
                    request.form.get("hospital_referencia") or ""
                ).strip(),
                eventualidades=(request.form.get("eventualidades") or "").strip(),
                created_by_id=current_user.id,
            )
            db.session.add(rec)
            db.session.commit()
            flash("Registro guardado correctamente.", "success")
            return redirect(url_for("listar"))

        except Exception as e:
            flash(f"Error guardando: {e}", "danger")

    return render_template("form.html")


@app.route("/editar/<int:rec_id>", methods=["GET", "POST"])
@login_required
@hospital_admin_required
def editar(rec_id):
    rec = EmergencyRecord.query.get_or_404(rec_id)

    if not current_user.is_admin and rec.hospital != current_user.hospital:
        flash("No tiene permiso para editar este registro.", "danger")
        return redirect(url_for("listar"))

    if request.method == "POST":
        try:
            fecha_str = request.form.get("fecha")
            nueva_fecha = parser.parse(fecha_str).date() if fecha_str else rec.fecha

            if current_user.is_admin:
                hospital_nombre = (request.form.get("hospital") or rec.hospital).strip()
                ok = Hospital.query.filter_by(
                    nombre=hospital_nombre, activo=True
                ).first()
                if not ok:
                    flash("Hospital inválido o inactivo.", "danger")
                    return render_template("edit.html", rec=rec)
                nuevo_hospital = hospital_nombre
            else:
                nuevo_hospital = current_user.hospital

            duplicado = EmergencyRecord.query.filter(
                EmergencyRecord.id != rec.id,
                EmergencyRecord.fecha == nueva_fecha,
                EmergencyRecord.hospital == nuevo_hospital,
            ).first()
            if duplicado:
                flash(
                    "Ya existe otro registro para este hospital en esa fecha.", "danger"
                )
                return render_template("edit.html", rec=rec)

            rec.fecha = nueva_fecha
            rec.hospital = nuevo_hospital

            def to_int(name, current):
                val = request.form.get(name, "")
                if val == "":
                    return current
                try:
                    return max(int(val), 0)
                except Exception:
                    return current

            rec.atenciones = to_int("atenciones", rec.atenciones)
            rec.ingresos = to_int("ingresos", rec.ingresos)
            rec.alta_voluntario = to_int("alta_voluntario", rec.alta_voluntario)
            rec.traslados = to_int("traslados", rec.traslados)
            rec.defunciones = to_int("defunciones", rec.defunciones)
            rec.motivo_traslado = (request.form.get("motivo_traslado") or "").strip()
            rec.hospital_referencia = (
                request.form.get("hospital_referencia") or ""
            ).strip()
            rec.eventualidades = (request.form.get("eventualidades") or "").strip()

            db.session.commit()
            flash("Registro actualizado correctamente.", "success")
            return redirect(url_for("listar"))

        except Exception as e:
            flash(f"Error actualizando: {e}", "danger")

    return render_template("edit.html", rec=rec)


@app.route("/eliminar/<int:rec_id>", methods=["POST"])
@login_required
@admin_required
def eliminar(rec_id):
    rec = EmergencyRecord.query.get_or_404(rec_id)
    db.session.delete(rec)
    db.session.commit()
    flash("Registro eliminado correctamente.", "success")
    return redirect(url_for("listar"))


@app.route("/listar")
@login_required
@hospital_admin_required
def listar():
    f_hospital = (request.args.get("hospital") or "").strip()
    f_desde = request.args.get("desde")
    f_hasta = request.args.get("hasta")

    # Paginación
    page = request.args.get("page", 1, type=int)
    if page < 1:
        page = 1
    per_page = 8  # puedes ajustar el tamaño de página

    q = EmergencyRecord.query

    # Filtro por hospital según rol
    if not current_user.is_admin:
        q = q.filter(EmergencyRecord.hospital == current_user.hospital)
    else:
        if f_hospital:
            q = q.filter(EmergencyRecord.hospital == f_hospital)

    # Parse de fechas
    def parse_date(s):
        try:
            return parser.parse(s).date()
        except Exception:
            return None

    d_desde = parse_date(f_desde) if f_desde else None
    d_hasta = parse_date(f_hasta) if f_hasta else None

    if d_desde:
        q = q.filter(EmergencyRecord.fecha >= d_desde)
    if d_hasta:
        q = q.filter(EmergencyRecord.fecha <= d_hasta)

    # Total de registros para la paginación
    total_registros = q.count()

    # Calcular número total de páginas
    pages = (total_registros + per_page - 1) // per_page if total_registros > 0 else 1
    if page > pages:
        page = pages

    # Traer solo la página actual
    registros = (
        q.order_by(EmergencyRecord.fecha.desc(), EmergencyRecord.id.desc())
        .offset((page - 1) * per_page)
        .limit(per_page)
        .all()
    )

    return render_template(
        "list.html",
        registros=registros,
        page=page,
        pages=pages,
        per_page=per_page,
        total_registros=total_registros,
    )


# ================== EXPORTAR CSV ==================
@app.route("/exportar_csv")
@login_required
@hospital_admin_required
def exportar_csv():
    f_hospital = (request.args.get("hospital") or "").strip()
    f_desde = request.args.get("desde")
    f_hasta = request.args.get("hasta")

    q = EmergencyRecord.query

    if not current_user.is_admin:
        q = q.filter(EmergencyRecord.hospital == current_user.hospital)
    else:
        if f_hospital:
            q = q.filter(EmergencyRecord.hospital == f_hospital)

    def parse_date(s):
        try:
            return parser.parse(s).date()
        except Exception:
            return None

    d_desde = parse_date(f_desde) if f_desde else None
    d_hasta = parse_date(f_hasta) if f_hasta else None

    if d_desde:
        q = q.filter(EmergencyRecord.fecha >= d_desde)
    if d_hasta:
        q = q.filter(EmergencyRecord.fecha <= d_hasta)

    registros = q.order_by(
        EmergencyRecord.fecha.desc(), EmergencyRecord.id.desc()
    ).all()

    si = StringIO()
    writer = csv.writer(si)
    writer.writerow(
        [
            "Fecha",
            "Hospital",
            "Atenciones",
            "Ingresos",
            "Alta Voluntario",
            "Traslados",
            "Motivo del traslado",
            "Hospital de referencia",
            "Defunciones",
            "Eventualidades",
        ]
    )
    for r in registros:
        writer.writerow(r.to_row())

    output = si.getvalue().encode("utf-8-sig")
    return Response(
        output,
        mimetype="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": "attachment; filename=registros_emergencias.csv"
        },
    )


# ================== EXPORTAR EXCEL ==================
@app.route("/exportar_excel")
@login_required
@hospital_admin_required
def exportar_excel():
    f_hospital = (request.args.get("hospital") or "").strip()
    f_desde = request.args.get("desde") or ""
    f_hasta = request.args.get("hasta") or ""

    def parse_date(s):
        try:
            return parser.parse(s).date() if s else None
        except Exception:
            return None

    d_desde = parse_date(f_desde)
    d_hasta = parse_date(f_hasta)

    if d_desde and d_hasta and d_desde > d_hasta:
        d_desde, d_hasta = d_hasta, d_desde
        f_desde, f_hasta = (
            d_desde.isoformat() if d_desde else "",
            d_hasta.isoformat() if d_hasta else "",
        )

    q = EmergencyRecord.query
    motivo_vacio = []

    if not current_user.is_admin:
        q = q.filter(EmergencyRecord.hospital == current_user.hospital)
        motivo_vacio.append(
            f"Rol usuario restringe a hospital '{current_user.hospital}'"
        )
    else:
        if f_hospital:
            q = q.filter(EmergencyRecord.hospital == f_hospital)
            motivo_vacio.append(f"Filtro hospital '{f_hospital}'")

    if d_desde:
        q = q.filter(EmergencyRecord.fecha >= d_desde)
        motivo_vacio.append(f"Desde {d_desde.isoformat()}")
    if d_hasta:
        q = q.filter(EmergencyRecord.fecha <= d_hasta)
        motivo_vacio.append(f"Hasta {d_hasta.isoformat()}")

    registros = q.order_by(EmergencyRecord.fecha.asc(), EmergencyRecord.id.asc()).all()

    reintento_sin_fechas = False
    if len(registros) == 0 and (d_desde or d_hasta):
        q2 = EmergencyRecord.query
        if not current_user.is_admin:
            q2 = q2.filter(EmergencyRecord.hospital == current_user.hospital)
        else:
            if f_hospital:
                q2 = q2.filter(EmergencyRecord.hospital == f_hospital)

        registros = q2.order_by(
            EmergencyRecord.fecha.asc(), EmergencyRecord.id.asc()
        ).all()
        reintento_sin_fechas = True

    wb = Workbook()
    ws = wb.active
    ws.title = "Registros"

    headers = [
        "Fecha",
        "Hospital",
        "Atenciones",
        "Ingresos",
        "Alta Voluntario",
        "Traslados",
        "Motivo del traslado",
        "Hospital de referencia",
        "Defunciones",
        "Eventualidades",
    ]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="0D6EFD")
    header_font = Font(color="FFFFFF", bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin = Side(border_style="thin", color="D0D7E2")
    border_all = Border(top=thin, left=thin, right=thin, bottom=thin)

    number_right = NamedStyle(name="number_right")
    number_right.number_format = "#,##0"
    number_right.alignment = Alignment(horizontal="right", vertical="center")

    date_center = NamedStyle(name="date_center")
    date_center.number_format = "yyyy-mm-dd"
    date_center.alignment = Alignment(horizontal="center", vertical="center")

    text_wrap = NamedStyle(name="text_wrap")
    text_wrap.alignment = Alignment(wrap_text=True, vertical="top")

    for st in (number_right, date_center, text_wrap):
        try:
            wb.add_named_style(st)
        except Exception:
            pass

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border_all

    row_start = 2
    for r in registros:
        ws.append(
            [
                r.fecha,
                r.hospital,
                r.atenciones,
                r.ingresos,
                r.alta_voluntario,
                r.traslados,
                r.motivo_traslado or "",
                r.hospital_referencia or "",
                r.defunciones,
                (r.eventualidades or "").replace("\r", " "),
            ]
        )

    COL_FECHA = 1
    COL_NUMS = [3, 4, 5, 6, 9]
    COL_TEXT_WRAP = [7, 8, 10]
    last_row = ws.max_row

    if last_row > 1:
        for row in range(row_start, last_row + 1):
            ws.cell(row=row, column=COL_FECHA).style = "date_center"
            for c in COL_NUMS:
                ws.cell(row=row, column=c).style = "number_right"
            for c in COL_TEXT_WRAP:
                ws.cell(row=row, column=c).style = "text_wrap"
            for c in range(1, len(headers) + 1):
                ws.cell(row=row, column=c).border = border_all

        total_row = last_row + 1
        ws.cell(row=total_row, column=1, value="Totales")
        ws.cell(row=total_row, column=1).font = Font(bold=True)
        ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="right")

        for c in COL_NUMS:
            col_letter = get_column_letter(c)
            ws.cell(
                row=total_row,
                column=c,
                value=f"=SUM({col_letter}{row_start}:{col_letter}{last_row})",
            ).style = "number_right"

        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=total_row, column=c)
            cell.border = border_all
            if c in COL_NUMS or c == 1:
                cell.fill = PatternFill("solid", fgColor="E9F2FF")

        last_row = total_row

    widths = {1: 12, 2: 38, 3: 12, 4: 12, 5: 16, 6: 12, 7: 28, 8: 28, 9: 12, 10: 50}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last_row}"

    summary = wb.create_sheet("Resumen", 0)
    summary["A1"] = "Exportación de Registros de Emergencias"
    summary["A1"].font = Font(size=14, bold=True)

    summary["A3"] = "Generado:"
    summary["B3"] = now_local().strftime("%Y-%m-%d %H:%M")
    summary["A4"] = "Hospital:"
    summary["B4"] = (
        f_hospital
        if (current_user.is_admin and f_hospital)
        else (current_user.hospital if not current_user.is_admin else "Todos")
    )
    summary["A5"] = "Desde:"
    summary["B5"] = f_desde or ""
    summary["A6"] = "Hasta:"
    summary["B6"] = f_hasta or ""
    summary["A7"] = "Registros exportados:"
    summary["B7"] = len(registros)

    summary["A9"] = "Notas:"
    notes = []
    if reintento_sin_fechas:
        notes.append("Sin resultados con fechas; se exportó sin filtros de fecha.")
    if motivo_vacio:
        notes.append("Filtros aplicados: " + "; ".join(motivo_vacio))
    summary["B9"] = "\n".join(notes) if notes else "—"

    summary.column_dimensions["A"].width = 20
    summary.column_dimensions["B"].width = 60
    for r in range(1, 11):
        for c in range(1, 3):
            summary.cell(row=r, column=c).alignment = Alignment(vertical="top")

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = "registros_emergencias.xlsx"
    return Response(
        bio.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ======================================================================
#   PDF HELPERS
# ======================================================================
def link_callback(uri, rel):
    """Resuelve rutas static para xhtml2pdf."""
    if uri.startswith("/static/"):
        uri = uri.lstrip("/")

    if uri.startswith("static/"):
        path = os.path.join(current_app.root_path, uri.replace("/", os.sep))
        if os.path.isfile(path):
            return path

    return uri


def render_pdf_from_html(html: str, pdf_filename: str = "reporte.pdf"):
    """Convierte HTML a PDF usando xhtml2pdf."""
    result = BytesIO()
    pisa_status = pisa.CreatePDF(
        src=html.encode("utf-8"),
        dest=result,
        encoding="utf-8",
        link_callback=link_callback,
    )

    if pisa_status.err:
        print("Errores xhtml2pdf:", pisa_status.err)
        return Response("Error generando PDF.", mimetype="text/plain")

    result.seek(0)
    response = make_response(result.read())
    response.headers["Content-Type"] = "application/pdf"
    response.headers["Content-Disposition"] = f'inline; filename="{pdf_filename}"'
    return response


# ======================================================================
#   DASHBOARD (Guardias + Internamientos)
# ======================================================================
import json
from datetime import (
    datetime as _dt_alias,
    date as _date_alias,
)  # para no chocar con helpers
from dateutil import parser as _parser_alias

from models import GuardiaCalendarioMensual  # 🆕 asegúrate de tener este import

import calendar as _calendar
from models import GuardiaCalendarioMensual, User


@app.route("/dashboard")
@login_required
def dashboard():
    f_hospital = (request.args.get("hospital") or "").strip()
    f_desde = request.args.get("desde") or ""
    f_hasta = request.args.get("hasta") or ""

    # ---------- Helper fechas ----------
    def parse_date(s):
        try:
            return _parser_alias.parse(s).date() if s else None
        except Exception:
            return None

    hoy = today_local()

    d_desde = parse_date(f_desde)
    d_hasta = parse_date(f_hasta)

    # Si no hay fechas, usar mes actual
    if not d_desde and not d_hasta:
        d_desde = hoy.replace(day=1)
        d_hasta = hoy
    else:
        if d_desde and not d_hasta:
            d_hasta = hoy
        if d_hasta and not d_desde:
            d_desde = d_hasta.replace(day=1)

    # Corregir si vienen invertidas
    if d_desde and d_hasta and d_desde > d_hasta:
        d_desde, d_hasta = d_hasta, d_desde

    # ---------- Alcance por rol ----------
    scope_hosp = user_hospital_scope()

    if scope_hosp:
        sel_hospital = scope_hosp
        f_hospital = scope_hosp
    else:
        sel_hospital = f_hospital or "Todos"

        # ============================================================
    #   CALENDARIO COMPLETO DE GUARDIAS (solo vista)
    # ============================================================
    anio_guardias = hoy.year
    mes_guardias = hoy.month
    hospital_guardias = scope_hosp or current_user.hospital

    # Médicos del hospital
    medicos_cal = (
        User.query.filter_by(hospital=hospital_guardias)
        .order_by(User.nombre.asc())
        .all()
    )

    # Paleta de colores por médico
    palette = [
        "#0d6efd",
        "#198754",
        "#ffc107",
        "#dc3545",
        "#20c997",
        "#6f42c1",
        "#fd7e14",
        "#0dcaf0",
        "#6c757d",
    ]
    color_map_cal = {}
    for idx, m in enumerate(medicos_cal):
        color_map_cal[m.id] = palette[idx % len(palette)]

    filas_cal = GuardiaCalendarioMensual.query.filter_by(
        hospital=hospital_guardias,
        anio=anio_guardias,
        mes=mes_guardias,
    ).all()

    asignaciones_cal = {f.medico_id: f for f in filas_cal}
    dias_por_medico_cal = {f.medico_id: f.dias_list for f in filas_cal}

    guardias_por_medico_cal = {
        m_id: len(dias) for m_id, dias in dias_por_medico_cal.items()
    }

    cal = _calendar.Calendar(firstweekday=0)
    semanas_cal = []
    for week in cal.monthdatescalendar(anio_guardias, mes_guardias):
        fila = []
        for d in week:
            if d.month != mes_guardias:
                fila.append(None)
                continue

            dia_num = d.day
            medicos_dia = []
            medicos_ids_dia = []
            for m in medicos_cal:
                lista = dias_por_medico_cal.get(m.id, [])
                if dia_num in lista:
                    medicos_dia.append(m)
                    medicos_ids_dia.append(m.id)

            fila.append(
                {
                    "fecha": d,
                    "medicos": medicos_dia,
                    "medicos_ids": medicos_ids_dia,
                }
            )
        semanas_cal.append(fila)


    # 🆕 ============================================================
    #   TUS GUARDIAS PARA ESTE MES (usuario actual)
    # ============================================================
    anio_guardias = hoy.year
    mes_guardias = hoy.month

    medico_id = current_user.id
    hospital_guardias = current_user.hospital  # en piloto es el hospital del usuario

    asign = GuardiaCalendarioMensual.query.filter_by(
        hospital=hospital_guardias,
        medico_id=medico_id,
        anio=anio_guardias,
        mes=mes_guardias,
    ).first()

    guardias_mes = asign.dias_list if asign else []  # lista de enteros [1, 7, 13, ...]

    # 🆕 Próxima guardia (dentro del mismo mes)
    proxima_guardia_fecha = None
    proxima_guardia_dias = None

    if guardias_mes:
        dias_ordenados = sorted(guardias_mes)
        futuros = [d for d in dias_ordenados if d >= hoy.day]

        if futuros:
            d_next = futuros[0]
            proxima_guardia_fecha = hoy.replace(day=d_next)
            proxima_guardia_dias = (proxima_guardia_fecha - hoy).days
        else:
            # Todas las guardias de este mes ya pasaron → no hay próxima
            proxima_guardia_fecha = None
            proxima_guardia_dias = None

    # ============================================================
    #   1) GUARDIAS DE EMERGENCIA
    # ============================================================
    qg = GuardiaEmergencia.query
    if f_hospital:
        qg = qg.filter(GuardiaEmergencia.hospital == f_hospital)

    if d_desde:
        qg = qg.filter(GuardiaEmergencia.fecha >= d_desde)
    if d_hasta:
        qg = qg.filter(GuardiaEmergencia.fecha <= d_hasta)

    guardias = qg.order_by(GuardiaEmergencia.fecha.asc()).all()

    # KPIs guardias
    guardia_total_pacientes = 0
    guardia_adultos = 0
    guardia_pediatricos = 0
    guardia_ginecologicas = 0
    guardia_fallecidos = 0
    guardia_referidos = 0
    guardia_ingresados_total = 0

    for g in guardias:
        guardia_total_pacientes += (
            (g.total_matutino or 0)
            + (g.total_vespertino or 0)
            + (g.total_nocturno or 0)
        )
        guardia_adultos += g.adultos or 0
        guardia_pediatricos += g.pediatricos or 0
        guardia_ginecologicas += g.ginecologicas or 0
        guardia_fallecidos += g.fallecidos or 0
        guardia_referidos += g.referidos or 0
        guardia_ingresados_total += g.ingresados_total or 0

    # ============================================================
    #   2) INTERNAMIENTOS
    # ============================================================
    qi = Internamiento.query
    if f_hospital:
        qi = qi.filter(Internamiento.hospital == f_hospital)

    if d_desde:
        qi = qi.filter(Internamiento.fecha >= d_desde)
    if d_hasta:
        qi = qi.filter(Internamiento.fecha <= d_hasta)

    internamientos_todos = qi.all()
    internamientos_activos = [i for i in internamientos_todos if not i.egresado]
    internamientos_egresados = [i for i in internamientos_todos if i.egresado]

    kpi_internamientos_activos = len(internamientos_activos)
    kpi_internamientos_egresados = len(internamientos_egresados)

    # ============================================================
    #   3) REGISTRO DIARIO DE EMERGENCIAS (EmergencyRecord)
    # ============================================================
    kpi_atenciones = 0
    kpi_ingresos = 0
    kpi_traslados = 0
    kpi_defunciones = 0
    labels = []
    chart_atenciones = []
    chart_ingresos = []
    chart_traslados = []
    chart_defunciones = []
    ranking = []

    referral_counts = {}
    motivo_counts = {}

    if current_user.is_admin or current_user.is_hospital_admin:
        qe = EmergencyRecord.query

        if f_hospital:
            qe = qe.filter(EmergencyRecord.hospital == f_hospital)
        elif scope_hosp:
            qe = qe.filter(EmergencyRecord.hospital == scope_hosp)

        if d_desde:
            qe = qe.filter(EmergencyRecord.fecha >= d_desde)
        if d_hasta:
            qe = qe.filter(EmergencyRecord.fecha <= d_hasta)

        registros = qe.order_by(EmergencyRecord.fecha.asc()).all()

        # KPIs de registro diario
        for r in registros:
            at = r.atenciones or 0
            ing = r.ingresos or 0
            tr = r.traslados or 0
            de = r.defunciones or 0

            kpi_atenciones += at
            kpi_ingresos += ing
            kpi_traslados += tr
            kpi_defunciones += de

            key = r.fecha.isoformat()
            if key not in labels:
                labels.append(key)

        series = {}
        for r in registros:
            key = r.fecha.isoformat()
            if key not in series:
                series[key] = {
                    "atenciones": 0,
                    "ingresos": 0,
                    "traslados": 0,
                    "defunciones": 0,
                }
            series[key]["atenciones"] += r.atenciones or 0
            series[key]["ingresos"] += r.ingresos or 0
            series[key]["traslados"] += r.traslados or 0
            series[key]["defunciones"] += r.defunciones or 0

            if (r.traslados or 0) > 0:
                hosp_ref = (r.hospital_referencia or "").strip()
                if hosp_ref:
                    referral_counts[hosp_ref] = referral_counts.get(hosp_ref, 0) + (
                        r.traslados or 0
                    )

                motivo = (r.motivo_traslado or "").strip()
                if motivo:
                    motivo_counts[motivo] = motivo_counts.get(motivo, 0) + (
                        r.traslados or 0
                    )

        labels = sorted(series.keys())
        chart_atenciones = [series[d]["atenciones"] for d in labels]
        chart_ingresos = [series[d]["ingresos"] for d in labels]
        chart_traslados = [series[d]["traslados"] for d in labels]
        chart_defunciones = [series[d]["defunciones"] for d in labels]

        if current_user.is_admin and not f_hospital:
            totales = {}
            for r in registros:
                totales.setdefault(r.hospital, 0)
                totales[r.hospital] += r.atenciones or 0
            ranking = sorted(
                ({"hospital": h, "atenciones": totales[h]} for h in totales),
                key=lambda x: x["atenciones"],
                reverse=True,
            )[:5]

    top_referrals = sorted(
        ({"hospital": h, "traslados": referral_counts[h]} for h in referral_counts),
        key=lambda x: x["traslados"],
        reverse=True,
    )[:6]

    referral_labels = [x["hospital"] for x in top_referrals]
    referral_values = [x["traslados"] for x in top_referrals]

    top_motivos = sorted(
        ({"motivo": m, "traslados": motivo_counts[m]} for m in motivo_counts),
        key=lambda x: x["traslados"],
        reverse=True,
    )[:6]

    motivo_labels = [x["motivo"] for x in top_motivos]
    motivo_values = [x["traslados"] for x in top_motivos]

    # 🆕 asegura que anio_actual / mes_actual existan
    anio_actual = hoy.year
    mes_actual = hoy.month

    return render_template(
        "dashboard.html",
        sel_hospital=sel_hospital,
        f_hospital=f_hospital,
        f_desde=d_desde.isoformat() if d_desde else "",
        f_hasta=d_hasta.isoformat() if d_hasta else "",
        guardia_total_pacientes=guardia_total_pacientes,
        guardia_adultos=guardia_adultos,
        guardia_pediatricos=guardia_pediatricos,
        guardia_ginecologicas=guardia_ginecologicas,
        guardia_fallecidos=guardia_fallecidos,
        guardia_referidos=guardia_referidos,
        guardia_ingresados_total=guardia_ingresados_total,
        kpi_internamientos_activos=kpi_internamientos_activos,
        kpi_internamientos_egresados=kpi_internamientos_egresados,
        kpi_atenciones=kpi_atenciones,
        kpi_ingresos=kpi_ingresos,
        kpi_traslados=kpi_traslados,
        kpi_defunciones=kpi_defunciones,
        labels=json.dumps(labels),
        data_atenciones=json.dumps(chart_atenciones),
        data_ingresos=json.dumps(chart_ingresos),
        data_traslados=json.dumps(chart_traslados),
        data_defunciones=json.dumps(chart_defunciones),
        ranking=ranking,
        referral_labels=json.dumps(referral_labels),
        referral_values=json.dumps(referral_values),
        motivo_labels=json.dumps(motivo_labels),
        motivo_values=json.dumps(motivo_values),
        anio_actual=anio_actual,
        mes_actual=mes_actual,
        guardias_mes=guardias_mes,                  # 🆕
        proxima_guardia_fecha=proxima_guardia_fecha,  # 🆕
        proxima_guardia_dias=proxima_guardia_dias,    # 🆕
        # 👇 datos para el calendario en el modal
        semanas_cal=semanas_cal,
        medicos_cal=medicos_cal,
        color_map_cal=color_map_cal,
        guardias_por_medico_cal=guardias_por_medico_cal,
        anio_guardias=anio_guardias,
        mes_guardias=mes_guardias,


    )


from sqlalchemy import func
from datetime import date as _date2, timedelta as _td2, datetime as _dt2
from dateutil import parser as _parser2
import uuid as _uuid2
import matplotlib.pyplot as _plt2
import numpy as _np2
import os as _os2
import logging

logger = logging.getLogger(__name__)


from sqlalchemy import func
from datetime import date, timedelta, datetime
from dateutil import parser
import uuid
import matplotlib.pyplot as plt
import os
import logging

logger = logging.getLogger(__name__)


def generar_grafica_dona_resumen(resumen_rows):
    """
    Gráfica de dona sin textos encima.
    Leyenda al lado con cuadrito de color + nombre + porcentaje.
    """
    if not resumen_rows:
        return None

    tmp_dir = _os2.path.join(app.static_folder, "img")
    _os2.makedirs(tmp_dir, exist_ok=True)

    labels = [row["hospital"] for row in resumen_rows]
    valores = [row["atenciones"] for row in resumen_rows]

    if all(v == 0 for v in valores):
        return None

    total_val = sum(valores)
    if total_val == 0:
        return None

    fig, ax = _plt2.subplots(figsize=(9, 5))

    # --- DONA ---
    wedges, _ = ax.pie(valores, startangle=90, wedgeprops=dict(width=0.35), radius=1.0)

    ax.set(aspect="equal")
    ax.set_title("ATENCIONES POR HOSPITAL", fontsize=12)

    # --- LEYENDA: nombre + porcentaje ---
    legend_labels = []
    for name, val in zip(labels, valores):
        pct = (val / total_val) * 100.0
        legend_labels.append(f"{name} — {pct:.1f}%")

    ax.legend(
        wedges,
        legend_labels,
        title="Hospitales",
        loc="center left",
        bbox_to_anchor=(1.05, 0.5),
        fontsize=9,
        title_fontsize=10,
        frameon=False,
    )

    filename = f"resumen_dona_{_uuid2.uuid4().hex}.png"
    filepath = _os2.path.join(tmp_dir, filename)

    fig.savefig(filepath, dpi=130, bbox_inches="tight")
    _plt2.close(fig)

    return f"/static/img/{filename}"


@app.route("/emergencias/resumen_pdf")
@login_required
@admin_required
def emergencias_resumen_pdf():
    """
    Resumen por hospital (tabla + dona) en PDF.
    """
    desde = None
    hasta = None

    desde_str = request.args.get("desde")
    hasta_str = request.args.get("hasta")

    if desde_str:
        try:
            desde = parser.parse(desde_str).date()
        except Exception:
            desde = None

    if hasta_str:
        try:
            hasta = parser.parse(hasta_str).date()
        except Exception:
            hasta = None

    if desde and not hasta:
        hasta = today_local()
    if hasta and not desde:
        desde = hasta.replace(day=1)

    if not desde or not hasta:
        try:
            anio = int(request.args.get("anio", now_local().year))
        except ValueError:
            anio = now_local().year

        try:
            mes = int(request.args.get("mes", now_local().month))
            if not (1 <= mes <= 12):
                mes = now_local().month
        except ValueError:
            mes = now_local().month

        desde = date(anio, mes, 1)
        if mes == 12:
            hasta = date(anio + 1, 1, 1) - timedelta(days=1)
        else:
            hasta = date(anio, mes + 1, 1) - timedelta(days=1)

    meses_es = [
        "ENERO",
        "FEBRERO",
        "MARZO",
        "ABRIL",
        "MAYO",
        "JUNIO",
        "JULIO",
        "AGOSTO",
        "SEPTIEMBRE",
        "OCTUBRE",
        "NOVIEMBRE",
        "DICIEMBRE",
    ]

    if (
        desde.day == 1
        and desde.month == hasta.month
        and desde.year == hasta.year
        and hasta.day >= 28
    ):
        mes_label = f"{meses_es[desde.month - 1]} {desde.year}"
    else:
        mes_label = f"Del {desde.strftime('%d/%m/%Y')} al {hasta.strftime('%d/%m/%Y')}"

    from models import EmergencyRecord

    base_q = (
        db.session.query(
            EmergencyRecord.hospital.label("hospital"),
            func.sum(EmergencyRecord.atenciones).label("atenciones"),
            func.sum(EmergencyRecord.ingresos).label("ingresos"),
            func.sum(EmergencyRecord.alta_voluntario).label("alta_voluntario"),
            func.sum(EmergencyRecord.traslados).label("traslados"),
            func.sum(EmergencyRecord.defunciones).label("defunciones"),
            func.count(
                func.nullif(
                    func.length(func.coalesce(EmergencyRecord.eventualidades, "")), 0
                )
            ).label("eventualidades_count"),
        )
        .filter(EmergencyRecord.fecha >= desde, EmergencyRecord.fecha <= hasta)
        .group_by(EmergencyRecord.hospital)
        .order_by(func.sum(EmergencyRecord.atenciones).desc())
    )

    resultados = base_q.all()

    resumen = []
    totales = {
        "atenciones": 0,
        "ingresos": 0,
        "alta_voluntario": 0,
        "traslados": 0,
        "defunciones": 0,
        "eventualidades_count": 0,
    }

    for row in resultados:
        data = {
            "hospital": row.hospital,
            "atenciones": int(row.atenciones or 0),
            "ingresos": int(row.ingresos or 0),
            "alta_voluntario": int(row.alta_voluntario or 0),
            "traslados": int(row.traslados or 0),
            "defunciones": int(row.defunciones or 0),
            "eventualidades_count": int(row.eventualidades_count or 0),
        }
        resumen.append(data)

        for k in (
            "atenciones",
            "ingresos",
            "alta_voluntario",
            "traslados",
            "defunciones",
        ):
            totales[k] += data[k]
        totales["eventualidades_count"] += data["eventualidades_count"]

    chart_url = generar_grafica_dona_resumen(resumen)

    html = render_template(
        "resumen_emergencias_pdf.html",
        mes_label=mes_label,
        desde=desde,
        hasta=hasta,
        resumen=resumen,
        totales=totales,
        chart_url=chart_url,
        generated_at=now_local(),
    )

    filename = (
        f"resumen_emergencias_{desde.strftime('%Y%m%d')}_{hasta.strftime('%Y%m%d')}.pdf"
    )
    response = render_pdf_from_html(html, pdf_filename=filename)

    if chart_url:
        fs_path = chart_url.replace("/static", app.static_folder)
        try:
            if os.path.exists(fs_path):
                os.remove(fs_path)
                app.logger.info(f"Imagen temporal eliminada: {fs_path}")
        except Exception as e:
            app.logger.warning(f"No se pudo eliminar imagen temporal {fs_path}: {e}")

    return response


# ======================================================================
#   CRUD HOSPITALES (ADMIN GENERAL)
# ======================================================================
@app.route("/hospitales")
@login_required
@admin_required
def hospitales_list():
    hospitales = Hospital.query.order_by(
        Hospital.activo.desc(), Hospital.nombre.asc()
    ).all()
    return render_template("hosp_list.html", hospitales=hospitales)


@app.route("/hospitales/nuevo", methods=["GET", "POST"])
@login_required
@admin_required
def hospitales_nuevo():
    if request.method == "POST":
        nombre = (request.form.get("nombre") or "").strip()
        logo_file = request.files.get("logo_file")

        if not nombre:
            flash("El nombre es obligatorio.", "danger")
            return render_template("hosp_form.html")

        if Hospital.query.filter_by(nombre=nombre).first():
            flash("Ya existe un hospital con ese nombre.", "danger")
            return render_template("hosp_form.html")

        logo_filename_db = None
        if logo_file and logo_file.filename:
            if not allowed_logo_file(logo_file.filename):
                flash("Tipo de logo no permitido. Use PNG/JPG/JPEG/GIF.", "danger")
                return render_template("hosp_form.html")

            safe_name = secure_filename(logo_file.filename)
            save_path = os.path.join(UPLOAD_FOLDER_LOGOS, safe_name)
            logo_file.save(save_path)
            logo_filename_db = os.path.join(LOGOS_SUBFOLDER, safe_name).replace(
                "\\", "/"
            )

        h = Hospital(nombre=nombre, activo=True, logo_filename=logo_filename_db)
        db.session.add(h)
        db.session.commit()
        flash("Hospital creado.", "success")
        return redirect(url_for("hospitales_list"))

    return render_template("hosp_form.html")


@app.route("/hospitales/editar/<int:h_id>", methods=["GET", "POST"])
@login_required
@admin_required
def hospitales_editar(h_id):
    h = Hospital.query.get_or_404(h_id)

    if request.method == "POST":
        nombre = (request.form.get("nombre") or "").strip()
        activo = True if request.form.get("activo") == "on" else False
        logo_file = request.files.get("logo_file")
        borrar_logo = True if request.form.get("borrar_logo") == "on" else False

        if not nombre:
            flash("El nombre es obligatorio.", "danger")
            return render_template("hosp_form.html", h=h)

        existe = Hospital.query.filter(
            Hospital.id != h.id, Hospital.nombre == nombre
        ).first()
        if existe:
            flash("Ya existe otro hospital con ese nombre.", "danger")
            return render_template("hosp_form.html", h=h)

        h.nombre = nombre
        h.activo = activo

        if borrar_logo:
            h.logo_filename = None

        if logo_file and logo_file.filename:
            if not allowed_logo_file(logo_file.filename):
                flash("Tipo de logo no permitido. Use PNG/JPG/JPEG/GIF.", "danger")
                return render_template("hosp_form.html", h=h)

            safe_name = secure_filename(logo_file.filename)
            save_path = os.path.join(UPLOAD_FOLDER_LOGOS, safe_name)
            logo_file.save(save_path)
            h.logo_filename = os.path.join(LOGOS_SUBFOLDER, safe_name).replace(
                "\\", "/"
            )

        db.session.commit()
        flash("Hospital actualizado.", "success")
        return redirect(url_for("hospitales_list"))

    return render_template("hosp_form.html", h=h)


@app.route("/hospitales/eliminar/<int:h_id>", methods=["POST"])
@login_required
@admin_required
def hospitales_eliminar(h_id):
    h = Hospital.query.get_or_404(h_id)
    h.activo = False
    db.session.commit()
    flash("Hospital desactivado (puedes reactivarlo editando).", "success")
    return redirect(url_for("hospitales_list"))


# ======================================================================
#   GESTIÓN DE USUARIOS
# ======================================================================
@app.route("/usuarios")
@login_required
@hospital_admin_required
def usuarios_list():
    q = User.query
    if current_user.is_admin:
        usuarios = q.order_by(
            User.is_admin.desc(), User.is_hospital_admin.desc(), User.username.asc()
        ).all()
    else:
        usuarios = (
            q.filter(User.hospital == current_user.hospital)
            .order_by(
                User.is_admin.desc(), User.is_hospital_admin.desc(), User.username.asc()
            )
            .all()
        )

    return render_template("users_list.html", usuarios=usuarios)


@app.route("/usuarios/nuevo", methods=["GET", "POST"])
@login_required
@hospital_admin_required
def usuarios_nuevo():
    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        hospital = (request.form.get("hospital") or "").strip()
        password1 = request.form.get("password1") or ""
        password2 = request.form.get("password2") or ""

        if not username:
            flash("El nombre de usuario es obligatorio.", "danger")
            return render_template("user_form.html")

        if User.query.filter_by(username=username).first():
            flash("Ya existe un usuario con ese nombre.", "danger")
            return render_template("user_form.html")

        if not password1:
            flash("La contraseña es obligatoria.", "danger")
            return render_template("user_form.html")

        if password1 != password2:
            flash("Las contraseñas no coinciden.", "danger")
            return render_template("user_form.html")

        if current_user.is_admin:
            hospital_final = hospital or None
            if hospital_final:
                h = Hospital.query.filter_by(nombre=hospital_final, activo=True).first()
                if not h:
                    flash("Hospital inválido o inactivo.", "danger")
                    return render_template("user_form.html")
        else:
            hospital_final = current_user.hospital
            if not hospital_final:
                flash("Tu usuario no tiene hospital asignado.", "danger")
                return render_template("user_form.html")

        if current_user.is_admin:
            is_admin = True if request.form.get("is_admin") == "on" else False
            is_hosp_admin = (
                True if request.form.get("is_hospital_admin") == "on" else False
            )
        else:
            is_admin = False
            is_hosp_admin = (
                True if request.form.get("is_hospital_admin") == "on" else False
            )

        u = User(
            username=username,
            hospital=hospital_final,
            is_admin=is_admin,
            is_hospital_admin=is_hosp_admin,
            nombre=request.form.get("nombre") or "",
            apellido=request.form.get("apellido") or "",
            cedula=request.form.get("cedula") or "",
            especialidad=request.form.get("especialidad") or "",
            cargo=request.form.get("cargo") or "",
            exequatur=request.form.get("exequatur") or "",
            telefono=request.form.get("telefono") or "",
            email=request.form.get("email") or "",
        )
        u.set_password(password1)
        db.session.add(u)
        db.session.commit()
        flash("Usuario creado correctamente.", "success")
        return redirect(url_for("usuarios_list"))

    return render_template("user_form.html")


@app.route("/usuarios/editar/<int:u_id>", methods=["GET", "POST"])
@login_required
@hospital_admin_required
def usuarios_editar(u_id):
    u = User.query.get_or_404(u_id)

    if not current_user.is_admin and u.hospital != current_user.hospital:
        flash("No puedes editar usuarios de otros hospitales.", "danger")
        return redirect(url_for("usuarios_list"))

    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        hospital = (request.form.get("hospital") or "").strip()
        password1 = request.form.get("password1") or ""
        password2 = request.form.get("password2") or ""

        if current_user.is_admin:
            is_admin = True if request.form.get("is_admin") == "on" else False
            is_hosp_admin = (
                True if request.form.get("is_hospital_admin") == "on" else False
            )
        else:
            is_admin = u.is_admin
            is_hosp_admin = (
                True if request.form.get("is_hospital_admin") == "on" else False
            )

        if not username:
            flash("El nombre de usuario es obligatorio.", "danger")
            return render_template("user_form.html", u=u)

        existe = User.query.filter(User.id != u.id, User.username == username).first()
        if existe:
            flash("Ya existe otro usuario con ese nombre.", "danger")
            return render_template("user_form.html", u=u)

        if current_user.is_admin:
            hospital_final = hospital or None
            if hospital_final:
                h = Hospital.query.filter_by(nombre=hospital_final, activo=True).first()
                if not h:
                    flash("Hospital inválido o inactivo.", "danger")
                    return render_template("user_form.html", u=u)
        else:
            hospital_final = current_user.hospital
            if not hospital_final:
                flash("Tu usuario no tiene hospital asignado.", "danger")
                return render_template("user_form.html", u=u)

        u.username = username
        u.hospital = hospital_final
        u.is_admin = is_admin
        u.is_hospital_admin = is_hosp_admin
        u.nombre = request.form.get("nombre") or u.nombre
        u.apellido = request.form.get("apellido") or u.apellido
        u.cedula = request.form.get("cedula") or u.cedula
        u.especialidad = request.form.get("especialidad") or u.especialidad
        u.cargo = request.form.get("cargo") or u.cargo
        u.exequatur = request.form.get("exequatur") or u.exequatur
        u.telefono = request.form.get("telefono") or u.telefono
        u.email = request.form.get("email") or u.email

        if password1 or password2:
            if password1 != password2:
                flash("Las contraseñas no coinciden.", "danger")
                return render_template("user_form.html", u=u)
            if not password1:
                flash("La nueva contraseña no puede estar vacía.", "danger")
                return render_template("user_form.html", u=u)
            u.set_password(password1)

        db.session.commit()
        flash("Usuario actualizado correctamente.", "success")
        return redirect(url_for("usuarios_list"))

    return render_template("user_form.html", u=u)


@app.route("/usuarios/eliminar/<int:u_id>", methods=["POST"])
@login_required
@hospital_admin_required
def usuarios_eliminar(u_id):
    u = User.query.get_or_404(u_id)

    if current_user.id == u.id:
        flash("No puedes eliminar tu propio usuario.", "danger")
        return redirect(url_for("usuarios_list"))

    if not current_user.is_admin and u.hospital != current_user.hospital:
        flash("No puedes eliminar usuarios de otros hospitales.", "danger")
        return redirect(url_for("usuarios_list"))

    db.session.delete(u)
    db.session.commit()
    flash("Usuario eliminado correctamente.", "success")
    return redirect(url_for("usuarios_list"))


# ======================================================================
#   GUARDIAS
# ======================================================================
@app.route("/guardias")
@login_required
def guardias_list():
    f_hospital = (request.args.get("hospital") or "").strip()
    f_desde = request.args.get("desde")
    f_hasta = request.args.get("hasta")

    def parse_date(s):
        try:
            return parser.parse(s).date()
        except Exception:
            return None

    d_desde = parse_date(f_desde) if f_desde else None
    d_hasta = parse_date(f_hasta) if f_hasta else None

    hoy = today_local()
    if not d_desde and not d_hasta:
        d_desde = hoy.replace(day=1)
        d_hasta = hoy
    else:
        if d_desde and not d_hasta:
            d_hasta = hoy
        if d_hasta and not d_desde:
            d_desde = d_hasta.replace(day=1)

    if d_desde and d_hasta and d_desde > d_hasta:
        d_desde, d_hasta = d_hasta, d_desde

    q = GuardiaEmergencia.query

    scope_hosp = user_hospital_scope()
    if scope_hosp:
        q = q.filter(GuardiaEmergencia.hospital == scope_hosp)
        hospitales_scope = [scope_hosp]
    else:
        if f_hospital:
            q = q.filter(GuardiaEmergencia.hospital == f_hospital)
            hospitales_scope = [f_hospital]
        else:
            hospitales_scope = [
                h.nombre for h in Hospital.query.filter_by(activo=True).all()
            ]

    if d_desde:
        q = q.filter(GuardiaEmergencia.fecha >= d_desde)
    if d_hasta:
        q = q.filter(GuardiaEmergencia.fecha <= d_hasta)

    guardias = q.order_by(
        GuardiaEmergencia.fecha.desc(), GuardiaEmergencia.id.desc()
    ).all()

    def date_range(start: date, end: date):
        cur = start
        while cur <= end:
            yield cur
            cur += timedelta(days=1)

    pendientes_por_hospital = {}
    total_pendientes = 0

    for hosp_name in hospitales_scope:
        fechas_con_guardia = {g.fecha for g in guardias if g.hospital == hosp_name}
        faltantes = [
            d for d in date_range(d_desde, d_hasta) if d not in fechas_con_guardia
        ]

        if faltantes:
            pendientes_por_hospital[hosp_name] = faltantes
            total_pendientes += len(faltantes)

    pendiente_count = 0
    pendiente_fechas = []
    if len(hospitales_scope) == 1:
        hname = hospitales_scope[0]
        if hname in pendientes_por_hospital:
            pendiente_fechas = [
                d.strftime("%Y-%m-%d") for d in pendientes_por_hospital[hname]
            ]
            pendiente_count = len(pendiente_fechas)

    return render_template(
        "guardias_list.html",
        guardias=guardias,
        rango_inicio=d_desde.strftime("%Y-%m-%d") if d_desde else "",
        rango_fin=d_hasta.strftime("%Y-%m-%d") if d_hasta else "",
        pendientes_por_hospital=pendientes_por_hospital,
        total_pendientes=total_pendientes,
        pendiente_count=pendiente_count,
        pendiente_fechas=pendiente_fechas,
    )


@app.route("/guardias/nuevo", methods=["GET", "POST"])
@login_required
def guardias_nuevo():
    if request.method == "POST":
        try:
            fecha_str = request.form.get("fecha")
            fecha = parser.parse(fecha_str).date() if fecha_str else today_local()

            hospital_nombre = (
                (request.form.get("hospital") or "").strip()
                if current_user.is_admin
                else current_user.hospital
            )

            if not hospital_nombre:
                flash("Debe indicar un hospital.", "danger")
                return render_template("guardias_form.html")

            existe = GuardiaEmergencia.query.filter_by(
                fecha=fecha, hospital=hospital_nombre
            ).first()
            if existe:
                flash(
                    "Ya existe una guardia registrada para esa fecha y hospital.",
                    "danger",
                )
                return redirect(url_for("guardias_editar", g_id=existe.id))

            def to_int(name):
                val = (request.form.get(name) or "0").strip()
                try:
                    return max(int(val), 0)
                except Exception:
                    return 0

            g = GuardiaEmergencia(
                fecha=fecha,
                hospital=hospital_nombre,
                medicos_emergencia=(
                    request.form.get("medicos_emergencia") or ""
                ).strip(),
                total_matutino=to_int("total_matutino"),
                total_vespertino=to_int("total_vespertino"),
                total_nocturno=to_int("total_nocturno"),
                adultos=to_int("adultos"),
                pediatricos=to_int("pediatricos"),
                ginecologicas=to_int("ginecologicas"),
                ingresados_total=to_int("ingresados_total"),
                ingresados_en_emergencia=to_int("ingresados_en_emergencia"),
                fallecidos=to_int("fallecidos"),
                traidos_911=to_int("traidos_911"),
                de_cuidados=to_int("de_cuidados"),
                referidos=to_int("referidos"),
                eventualidades=(request.form.get("eventualidades") or "").strip(),
                firma=(request.form.get("firma") or "").strip(),
                created_by_id=current_user.id,
            )

            db.session.add(g)
            db.session.commit()
            flash("Guardia creada correctamente.", "success")
            return redirect(url_for("guardias_list"))

        except Exception as e:
            flash(f"Error guardando guardia: {e}", "danger")

    hoy = now_local().strftime("%Y-%m-%d")
    return render_template("guardias_form.html", hoy=hoy, g=None)


@app.route("/guardias/editar/<int:g_id>", methods=["GET", "POST"])
@login_required
def guardias_editar(g_id):
    g = GuardiaEmergencia.query.get_or_404(g_id)

    scope_hosp = user_hospital_scope()
    if scope_hosp and g.hospital != scope_hosp:
        flash("No puedes editar guardias de otros hospitales.", "danger")
        return redirect(url_for("guardias_list"))

    if request.method == "POST":
        try:
            fecha_str = request.form.get("fecha")
            nueva_fecha = parser.parse(fecha_str).date() if fecha_str else g.fecha

            hospital_nombre = (
                (request.form.get("hospital") or g.hospital).strip()
                if current_user.is_admin
                else g.hospital
            )

            dup = GuardiaEmergencia.query.filter(
                GuardiaEmergencia.id != g.id,
                GuardiaEmergencia.fecha == nueva_fecha,
                GuardiaEmergencia.hospital == hospital_nombre,
            ).first()
            if dup:
                flash("Ya existe otra guardia para esa fecha y hospital.", "danger")
                return render_template("guardias_form.html", g=g)

            def to_int(name, current):
                val = request.form.get(name, "")
                if val == "":
                    return current
                try:
                    return max(int(val), 0)
                except Exception:
                    return current

            g.fecha = nueva_fecha
            g.hospital = hospital_nombre
            g.medicos_emergencia = (
                request.form.get("medicos_emergencia") or ""
            ).strip()
            g.total_matutino = to_int("total_matutino", g.total_matutino)
            g.total_vespertino = to_int("total_vespertino", g.total_vespertino)
            g.total_nocturno = to_int("total_nocturno", g.total_nocturno)
            g.adultos = to_int("adultos", g.adultos)
            g.pediatricos = to_int("pediatricos", g.pediatricos)
            g.ginecologicas = to_int("ginecologicas", g.ginecologicas)
            g.ingresados_total = to_int("ingresados_total", g.ingresados_total)
            g.ingresados_en_emergencia = to_int(
                "ingresados_en_emergencia", g.ingresados_en_emergencia
            )
            g.fallecidos = to_int("fallecidos", g.fallecidos)
            g.traidos_911 = to_int("traidos_911", g.traidos_911)
            g.de_cuidados = to_int("de_cuidados", g.de_cuidados)
            g.referidos = to_int("referidos", g.referidos)
            g.eventualidades = (request.form.get("eventualidades") or "").strip()
            g.firma = (request.form.get("firma") or "").strip()

            db.session.commit()
            flash("Guardia actualizada correctamente.", "success")
            return redirect(url_for("guardias_list"))

        except Exception as e:
            flash(f"Error actualizando guardia: {e}", "danger")

    hoy = now_local().strftime("%Y-%m-%d")
    return render_template("guardias_edit.html", hoy=hoy, g=g)


@app.route("/guardias/eliminar/<int:g_id>", methods=["POST"])
@login_required
@hospital_admin_required
def guardias_eliminar(g_id):
    g = GuardiaEmergencia.query.get_or_404(g_id)

    scope_hosp = user_hospital_scope()
    if scope_hosp and g.hospital != scope_hosp:
        flash("No puedes eliminar guardias de otros hospitales.", "danger")
        return redirect(url_for("guardias_list"))

    db.session.delete(g)
    db.session.commit()
    flash("Guardia eliminada correctamente.", "success")
    return redirect(url_for("guardias_list"))


@app.route("/guardias/<int:g_id>/pdf")
@login_required
def guardias_pdf(g_id):
    g = GuardiaEmergencia.query.get_or_404(g_id)

    hosp = Hospital.query.filter_by(nombre=g.hospital, activo=True).first()

    scope_hosp = user_hospital_scope()
    if scope_hosp and g.hospital != scope_hosp:
        flash("No puedes ver guardias de otros hospitales.", "danger")
        return redirect(url_for("guardias_list"))

    html = render_template("guardia_pdf.html", g=g, hosp=hosp, generated_at=now_local())
    filename = f"guardia_{g.hospital.replace(' ', '_')}_{g.fecha.isoformat()}.pdf"
    return render_pdf_from_html(html, pdf_filename=filename)


# ======================================================================
#   Trar Datos de la Guardia a Reporte Diario
# ======================================================================
@app.route("/api/guardia_por_dia", methods=["GET"])
@login_required
def api_guardia_por_dia():
    fecha_str = request.args.get("fecha", "").strip()
    if not fecha_str:
        return jsonify({"ok": False, "error": "Falta la fecha."}), 400

    try:
        fecha = parser.parse(fecha_str).date()
    except Exception:
        return jsonify({"ok": False, "error": "Fecha inválida."}), 400

    if current_user.is_admin:
        hospital = (request.args.get("hospital") or "").strip()
        if not hospital:
            return (
                jsonify({"ok": False, "error": "Falta seleccionar el hospital."}),
                400,
            )
    else:
        hospital = current_user.hospital

    if not hospital:
        return (
            jsonify({"ok": False, "error": "El usuario no tiene hospital asignado."}),
            400,
        )

    guardia = GuardiaEmergencia.query.filter_by(fecha=fecha, hospital=hospital).first()

    if not guardia:
        return (
            jsonify(
                {
                    "ok": True,
                    "found": False,
                    "message": "No hay guardia registrada para esa fecha en este hospital.",
                }
            ),
            200,
        )

    total_pacientes = (
        (guardia.total_matutino or 0)
        + (guardia.total_vespertino or 0)
        + (guardia.total_nocturno or 0)
    )

    data = {
        "ok": True,
        "found": True,
        "atenciones": total_pacientes,
        "ingresos": guardia.ingresados_total or 0,
        "traslados": guardia.referidos or 0,
        "defunciones": guardia.fallecidos or 0,
        "eventualidades": guardia.eventualidades or "",
    }
    return jsonify(data), 200


# ======================================================================
#   INTERNAMIENTOS
# ======================================================================
@app.route("/internamientos")
@login_required
def internamientos_list():
    f_hospital = (request.args.get("hospital") or "").strip()
    f_desde = request.args.get("desde")
    f_hasta = request.args.get("hasta")

    mostrar_egresados = request.args.get("egresados") == "1"

    q = Internamiento.query.options(joinedload(Internamiento.created_by))

    # Filtrar por egresados o no
    if not mostrar_egresados:
        q = q.filter(Internamiento.egresado == False)

    # Alcance por hospital
    scope_hosp = user_hospital_scope()
    if scope_hosp:
        q = q.filter(Internamiento.hospital == scope_hosp)
    else:
        if f_hospital:
            q = q.filter(Internamiento.hospital == f_hospital)

    # ---------------- Filtro de fechas (solo si se usa) ----------------
    def parse_date(s):
        try:
            return parser.parse(s).date()
        except Exception:
            return None

    d_desde = parse_date(f_desde) if f_desde else None
    d_hasta = parse_date(f_hasta) if f_hasta else None

    hoy = today_local()
    aplicar_filtro_fecha = bool(d_desde or d_hasta)

    if aplicar_filtro_fecha:
        if d_desde and not d_hasta:
            d_hasta = hoy
        if d_hasta and not d_desde:
            d_desde = d_hasta.replace(day=1)

        if d_desde and d_hasta and d_desde > d_hasta:
            d_desde, d_hasta = d_hasta, d_desde

        if d_desde:
            q = q.filter(Internamiento.fecha >= d_desde)
        if d_hasta:
            q = q.filter(Internamiento.fecha <= d_hasta)

    # ---------------- ORDENADO FINAL ----------------
    if current_user.is_admin and not scope_hosp and not f_hospital:
        # Admin general viendo toda la red → agrupar por hospital
        internamientos = q.order_by(
            Internamiento.hospital.asc(),
            Internamiento.area.asc(),
            Internamiento.habitacion.asc(),
            Internamiento.nombre_paciente.asc(),
        ).all()
    else:
        internamientos = q.order_by(
            Internamiento.fecha.desc(), Internamiento.id.desc()
        ).all()

    # --------- Totales por hospital (número de pacientes) ---------
    totales_por_hospital = {}
    for i in internamientos:
        totales_por_hospital[i.hospital] = totales_por_hospital.get(i.hospital, 0) + 1

    return render_template(
        "internamientos_list.html",
        internamientos=internamientos,
        rango_inicio=(
            d_desde.strftime("%Y-%m-%d") if aplicar_filtro_fecha and d_desde else ""
        ),
        rango_fin=(
            d_hasta.strftime("%Y-%m-%d") if aplicar_filtro_fecha and d_hasta else ""
        ),
        f_hospital=f_hospital,
        mostrar_egresados=mostrar_egresados,
        totales_por_hospital=totales_por_hospital,  # 👈 NUEVO
    )


@app.route("/internamientos/nuevo", methods=["GET", "POST"])
@login_required
def internamientos_nuevo():
    if request.method == "POST":
        try:
            fecha_str = request.form.get("fecha")
            fecha = parser.parse(fecha_str).date() if fecha_str else today_local()

            if current_user.is_admin or current_user.is_hospital_admin:
                hospital_nombre = (
                    request.form.get("hospital") or current_user.hospital or ""
                ).strip()
            else:
                hospital_nombre = current_user.hospital

            if not hospital_nombre:
                flash("Debe indicar un hospital.", "danger")
                return render_template("internamientos_form.html", ir=None)

            def to_int_or_none(name):
                val = (request.form.get(name) or "").strip()
                if not val:
                    return None
                try:
                    return int(val)
                except Exception:
                    return None

            dia_ingreso_str = (request.form.get("dia_ingreso") or "").strip()
            dia_ingreso_val = None
            if dia_ingreso_str:
                try:
                    dia_ingreso_val = parser.parse(dia_ingreso_str).date()
                except Exception:
                    dia_ingreso_val = None

            ir = Internamiento(
                fecha=fecha,
                hospital=hospital_nombre,
                area=(request.form.get("area") or "").strip(),
                habitacion=(request.form.get("habitacion") or "").strip(),
                nombre_paciente=(request.form.get("nombre_paciente") or "").strip(),
                edad=to_int_or_none("edad"),
                signos_vitales=(request.form.get("signos_vitales") or "").strip(),
                diagnosticos=(request.form.get("diagnosticos") or "").strip(),
                condicion_plan=(request.form.get("condicion_plan") or "").strip(),
                origen_ingreso=(request.form.get("origen_ingreso") or "").strip(),
                observaciones=(request.form.get("observaciones") or "").strip(),
                egresado=True if request.form.get("egresado") == "on" else False,
                dia_ingreso=dia_ingreso_val,
                created_by_id=current_user.id,
            )

            db.session.add(ir)
            db.session.commit()
            flash("Internamiento registrado correctamente.", "success")
            return redirect(url_for("internamientos_list"))

        except Exception as e:
            flash(f"Error guardando internamiento: {e}", "danger")

    hoy = now_local().strftime("%Y-%m-%d")
    return render_template("internamientos_form.html", hoy=hoy, ir=None)


@app.route("/internamientos/editar/<int:i_id>", methods=["GET", "POST"])
@login_required
def internamientos_editar(i_id):
    ir = Internamiento.query.get_or_404(i_id)

    scope_hosp = user_hospital_scope()
    if scope_hosp and ir.hospital != scope_hosp:
        flash("No puedes editar internamientos de otros hospitales.", "danger")
        return redirect(url_for("internamientos_list"))

    if request.method == "POST":
        try:
            fecha_str = request.form.get("fecha")
            ir.fecha = parser.parse(fecha_str).date() if fecha_str else ir.fecha

            if current_user.is_admin or current_user.is_hospital_admin:
                hospital_nombre = (request.form.get("hospital") or ir.hospital).strip()
                ir.hospital = hospital_nombre

            def to_int_or_none(name, current=None):
                val = (request.form.get(name) or "").strip()
                if val == "":
                    return current
                try:
                    return int(val)
                except Exception:
                    return current

            ir.edad = to_int_or_none("edad", ir.edad)

            dia_ingreso_str = (request.form.get("dia_ingreso") or "").strip()
            if dia_ingreso_str:
                try:
                    ir.dia_ingreso = parser.parse(dia_ingreso_str).date()
                except Exception:
                    pass

            ir.area = (request.form.get("area") or "").strip()
            ir.habitacion = (request.form.get("habitacion") or "").strip()
            ir.nombre_paciente = (request.form.get("nombre_paciente") or "").strip()
            ir.signos_vitales = (request.form.get("signos_vitales") or "").strip()
            ir.diagnosticos = (request.form.get("diagnosticos") or "").strip()
            ir.condicion_plan = (request.form.get("condicion_plan") or "").strip()
            ir.origen_ingreso = (request.form.get("origen_ingreso") or "").strip()
            ir.observaciones = (request.form.get("observaciones") or "").strip()
            ir.egresado = True if request.form.get("egresado") == "on" else False

            db.session.commit()
            flash("Internamiento actualizado correctamente.", "success")
            return redirect(url_for("internamientos_list"))

        except Exception as e:
            flash(f"Error actualizando internamiento: {e}", "danger")

    hoy = now_local().strftime("%Y-%m-%d")
    return render_template("internamientos_form.html", hoy=hoy, ir=ir)


@app.route("/internamientos/eliminar/<int:i_id>", methods=["POST"])
@login_required
def internamientos_eliminar(i_id):
    ir = Internamiento.query.get_or_404(i_id)

    scope_hosp = user_hospital_scope()
    if scope_hosp and ir.hospital != scope_hosp:
        flash("No puedes eliminar internamientos de otros hospitales.", "danger")
        return redirect(url_for("internamientos_list"))

    db.session.delete(ir)
    db.session.commit()
    flash("Internamiento eliminado correctamente.", "success")
    return redirect(url_for("internamientos_list"))


@app.route("/internamientos/pdf")
@login_required
def internamientos_pdf():
    # --------- Helper para resolver hospital según rol/filtro ----------
    scope_hosp = user_hospital_scope()  # None para admin general
    hospital_param = (request.args.get("hospital") or "").strip()

    # scope_hosp:
    #   - None  → admin general
    #   - "Hospital X" → admin hospital / usuario normal
    if scope_hosp:
        hospital_nombre = scope_hosp
        multi_hospital = False
    else:
        # Admin general: usa el hospital del filtro si viene, si no → todos
        hospital_nombre = hospital_param or None
        multi_hospital = hospital_nombre is None

    # 1) SOLO VALIDACIÓN
    if request.args.get("validar") == "1":
        fecha_str = request.args.get("fecha")
        try:
            fecha_reporte = (
                parser.parse(fecha_str).date() if fecha_str else today_local()
            )
        except Exception:
            fecha_reporte = today_local()

        # Internamientos activos hasta la fecha del reporte
        q_val = Internamiento.query.filter(
            Internamiento.egresado == False, Internamiento.fecha <= fecha_reporte
        )

        # Si hay hospital fijo (usuario normal o admin con filtro) se filtra
        if hospital_nombre:
            q_val = q_val.filter(Internamiento.hospital == hospital_nombre)

        activos = q_val.all()

        faltantes = []
        for i in activos:
            ultima_actualizacion = (
                i.fecha_actualizacion.date() if i.fecha_actualizacion else None
            )
            if not ultima_actualizacion or ultima_actualizacion != fecha_reporte:
                item = {
                    "id": i.id,
                    "nombre": i.nombre_paciente,
                    "area": i.area or "",
                    "habitacion": i.habitacion or "",
                    "ultima_actualizacion": (
                        ultima_actualizacion.isoformat()
                        if ultima_actualizacion
                        else None
                    ),
                }
                # Para admins, puede ser útil saber el hospital si es multi
                if multi_hospital:
                    item["hospital"] = i.hospital
                faltantes.append(item)

        # Para admin general y admin de hospital NO se bloquea
        if current_user.is_admin or current_user.is_hospital_admin:
            ok = True
        else:
            ok = len(faltantes) == 0

        return jsonify({"ok": ok, "faltantes": faltantes})

    # 2) GENERAR PDF
    if request.args.get("generar") == "1":
        fecha_str = request.args.get("fecha")
        try:
            fecha_reporte = (
                parser.parse(fecha_str).date() if fecha_str else today_local()
            )
        except Exception:
            fecha_reporte = today_local()

        mostrar_egresados = request.args.get("egresados") == "1"

        q = Internamiento.query.options(joinedload(Internamiento.created_by))

        # Filtro por hospital (solo si hay uno definido).
        # Si multi_hospital=True (admin general sin filtro), NO se filtra.
        if hospital_nombre:
            q = q.filter(Internamiento.hospital == hospital_nombre)

        if not mostrar_egresados:
            q = q.filter(Internamiento.egresado == False)

        q = q.filter(Internamiento.fecha <= fecha_reporte)

        # Orden:
        #   - multi_hospital → agrupar por hospital
        #   - un solo hospital → solo por área/habitación/nombre
        if multi_hospital:
            internamientos = q.order_by(
                Internamiento.hospital.asc(),
                Internamiento.area.asc(),
                Internamiento.habitacion.asc(),
                Internamiento.nombre_paciente.asc(),
            ).all()
        else:
            internamientos = q.order_by(
                Internamiento.area.asc(),
                Internamiento.habitacion.asc(),
                Internamiento.nombre_paciente.asc(),
            ).all()

        # ================== TOTALES POR HOSPITAL ==================
        totales_por_hospital = {}
        for i in internamientos:
            totales_por_hospital[i.hospital] = (
                totales_por_hospital.get(i.hospital, 0) + 1
            )
        # ==========================================================

        # ================== NOTA AUTOMÁTICA DE NO ACTUALIZADOS ==================
        q_val = Internamiento.query.filter(
            Internamiento.egresado == False, Internamiento.fecha <= fecha_reporte
        )
        if hospital_nombre:
            q_val = q_val.filter(Internamiento.hospital == hospital_nombre)

        activos = q_val.all()

        notas_auto = []
        for i in activos:
            ultima_actualizacion = (
                i.fecha_actualizacion.date() if i.fecha_actualizacion else None
            )
            if not ultima_actualizacion or ultima_actualizacion != fecha_reporte:
                if ultima_actualizacion:
                    fecha_txt = ultima_actualizacion.strftime("%d/%m/%Y")
                else:
                    fecha_txt = "Sin registro de actualización"

                # Un paciente por línea
                if multi_hospital:
                    # Incluimos hospital cuando el reporte es de toda la red
                    notas_auto.append(
                        f"- {i.nombre_paciente} | {i.hospital} | "
                        f"Área: {i.area or ''}, Hab.: {i.habitacion or ''} "
                        f"– Última actualización: {fecha_txt}"
                    )
                else:
                    notas_auto.append(
                        f"- {i.nombre_paciente} "
                        f"(Área: {i.area or ''}, Hab.: {i.habitacion or ''}) "
                        f"– Última actualización: {fecha_txt}"
                    )

        observaciones_generales = (
            request.args.get("observaciones_generales") or ""
        ).strip()

        if notas_auto:
            bloque = (
                "Pacientes con internamiento activo sin actualización "
                f"en la fecha del reporte ({fecha_reporte.strftime('%d/%m/%Y')}):\n"
                + "\n".join(notas_auto)
            )
            if observaciones_generales:
                observaciones_generales = (
                    observaciones_generales.rstrip() + "\n\n\n" + bloque
                )
            else:
                observaciones_generales = bloque
        # =======================================================================

        # Si hay hospital específico, buscamos su registro; si no, reporte general
        hosp = None
        if hospital_nombre and not multi_hospital:
            hosp = Hospital.query.filter_by(nombre=hospital_nombre, activo=True).first()

        html = render_template(
            "internamientos_pdf.html",
            internamientos=internamientos,
            hosp=hosp,
            hospital_nombre=hospital_nombre,
            multi_hospital=multi_hospital,
            totales_por_hospital=totales_por_hospital,
            fecha_reporte=fecha_reporte,
            generated_at=now_local(),
            mostrar_egresados=mostrar_egresados,
            observaciones_generales=observaciones_generales,
        )

        safe_name = (hospital_nombre or "Todos_los_hospitales").replace(" ", "_")
        filename = f"internamientos_{safe_name}_{fecha_reporte.isoformat()}.pdf"
        return render_pdf_from_html(html, pdf_filename=filename)

    return jsonify({"error": "Modo inválido"}), 400


# =======================================================================
# Calendario de Guardias
# ======================================================================
import calendar
from models import GuardiaCalendarioMensual, User

from datetime import date
import calendar
from models import GuardiaCalendarioMensual, User


@app.route("/guardias/calendario")
@login_required
@hospital_admin_required
def guardias_calendario():
    try:
        anio = int(request.args.get("anio", now_local().year))
    except ValueError:
        anio = now_local().year

    try:
        mes = int(request.args.get("mes", now_local().month))
        if not (1 <= mes <= 12):
            mes = now_local().month
    except ValueError:
        mes = now_local().month

    hospital = current_user.hospital

    medicos = User.query.filter_by(hospital=hospital).order_by(User.nombre.asc()).all()

    # 🎨 Paleta de colores
    palette = [
        "#0d6efd",
        "#198754",
        "#ffc107",
        "#dc3545",
        "#20c997",
        "#6f42c1",
        "#fd7e14",
        "#0dcaf0",
        "#6c757d",
    ]
    color_map = {}
    for idx, m in enumerate(medicos):
        color_map[m.id] = palette[idx % len(palette)]

    filas = GuardiaCalendarioMensual.query.filter_by(
        hospital=hospital, anio=anio, mes=mes
    ).all()

    asignaciones = {f.medico_id: f for f in filas}
    dias_por_medico = {f.medico_id: f.dias_list for f in filas}

    # 👉 Conteo de días por médico para la leyenda
    guardias_por_medico = {m_id: len(dias) for m_id, dias in dias_por_medico.items()}

    cal = calendar.Calendar(firstweekday=0)
    semanas = []

    hoy = now_local().date()  # para resaltar hoy

    for week in cal.monthdatescalendar(anio, mes):
        fila = []
        for d in week:
            if d.month != mes:
                fila.append(None)
                continue

            dia_num = d.day
            medicos_dia = []
            medicos_ids_dia = []
            for m in medicos:
                lista = dias_por_medico.get(m.id, [])
                if dia_num in lista:
                    medicos_dia.append(m)
                    medicos_ids_dia.append(m.id)

            fila.append(
                {
                    "fecha": d,
                    "medicos": medicos_dia,
                    "medicos_ids": medicos_ids_dia,
                    "es_hoy": (d == hoy),
                }
            )
        semanas.append(fila)

    return render_template(
        "guardias_calendario.html",
        semanas=semanas,
        medicos=medicos,
        asignaciones=asignaciones,
        color_map=color_map,
        guardias_por_medico=guardias_por_medico,
        anio=anio,
        mes=mes,
    )


@app.route("/guardias/calendario/guardar-dia", methods=["POST"])
@login_required
@hospital_admin_required
def guardar_guardias_dia():
    hospital = current_user.hospital

    try:
        anio = int(request.form.get("anio", now_local().year))
    except ValueError:
        anio = now_local().year

    try:
        mes = int(request.form.get("mes", now_local().month))
    except ValueError:
        mes = now_local().month

    try:
        dia = int(request.form.get("dia"))
    except (TypeError, ValueError):
        flash("Día inválido.", "danger")
        return redirect(url_for("guardias_calendario", anio=anio, mes=mes))

    # IDs de médicos seleccionados para ese día
    ids_str = request.form.getlist("medicos_ids")
    try:
        ids_seleccionados = {int(x) for x in ids_str}
    except ValueError:
        ids_seleccionados = set()

    medicos = User.query.filter_by(hospital=hospital).all()

    for m in medicos:
        asign = GuardiaCalendarioMensual.query.filter_by(
            hospital=hospital, medico_id=m.id, anio=anio, mes=mes
        ).first()

        if m.id in ids_seleccionados:
            # Debe tener el día en su lista
            if not asign:
                asign = GuardiaCalendarioMensual(
                    hospital=hospital, medico_id=m.id, anio=anio, mes=mes
                )
                db.session.add(asign)

            lista = asign.dias_list
            if dia not in lista:
                lista.append(dia)
            asign.dias_list = lista

        else:
            # NO debe tener este día
            if asign:
                lista = asign.dias_list
                if dia in lista:
                    lista = [d for d in lista if d != dia]
                    if lista:
                        asign.dias_list = lista
                    else:
                        db.session.delete(asign)

    db.session.commit()
    flash(f"Guardias del día {dia} actualizadas.", "success")
    return redirect(url_for("guardias_calendario", anio=anio, mes=mes))


@app.route("/guardias/calendario/guardar-medicos", methods=["POST"])
@login_required
@hospital_admin_required
def guardar_guardias_medicos():
    hospital = current_user.hospital

    try:
        anio = int(request.form.get("anio", now_local().year))
    except ValueError:
        anio = now_local().year

    try:
        mes = int(request.form.get("mes", now_local().month))
    except ValueError:
        mes = now_local().month

    medicos = User.query.filter_by(hospital=hospital).order_by(User.nombre.asc()).all()

    for m in medicos:
        key = f"dias_{m.id}"
        dias_str = (request.form.get(key) or "").strip()

        # buscamos si ya existe registro ese mes para ese médico
        asign = GuardiaCalendarioMensual.query.filter_by(
            hospital=hospital, medico_id=m.id, anio=anio, mes=mes
        ).first()

        if not dias_str:
            # si no escribió nada y existía, borrar asignación
            if asign:
                db.session.delete(asign)
            continue

        # parsear los días "1, 7, 13"
        try:
            lista_dias = [
                int(x)
                for x in dias_str.replace(" ", "").split(",")
                if x.strip().isdigit()
            ]
        except ValueError:
            lista_dias = []

        if not lista_dias:
            # nada válido
            if asign:
                db.session.delete(asign)
            continue

        if not asign:
            asign = GuardiaCalendarioMensual(
                hospital=hospital, medico_id=m.id, anio=anio, mes=mes
            )
            db.session.add(asign)

        asign.dias_list = lista_dias  # usa el setter para formatear

    db.session.commit()
    flash("Programación de guardias por médico guardada.", "success")
    return redirect(url_for("guardias_calendario", anio=anio, mes=mes))


#================================================================================
# PDF GUARDIAS POR MÉDICO (días de guardia)
#================================================================================
import calendar
from flask import Response
from models import GuardiaCalendarioMensual, User  # Hospital lo manejamos dentro con try/except

@app.route("/guardias/calendario_pdf")
@login_required
def guardias_calendario_pdf():
    """
    PDF institucional de calendario de guardias:
    Tabla por médico: nombre, días de guardia y total de guardias.
    Solo se muestran médicos con guardias asignadas.
    Ordenados por el primer día de guardia.
    """
    hoy = today_local()

    # Año y mes
    try:
        anio = int(request.args.get("anio", hoy.year))
    except ValueError:
        anio = hoy.year

    try:
        mes = int(request.args.get("mes", hoy.month))
        if not (1 <= mes <= 12):
            mes = hoy.month
    except ValueError:
        mes = hoy.month

    # Hospital según scope / usuario
    scope_hosp = user_hospital_scope()  # mismo helper que usas en dashboard
    if scope_hosp:
        hospital = scope_hosp
    else:
        hospital = current_user.hospital

    # Intentar obtener objeto Hospital para el logo (si existe el modelo)
    hosp = None
    try:
        from models import Hospital
        hosp = Hospital.query.filter_by(nombre=hospital).first()
    except Exception:
        hosp = None

    # Médicos de ese hospital
    medicos = (
        User.query
        .filter_by(hospital=hospital)
        .order_by(User.nombre.asc())
        .all()
    )

    # Asignaciones de guardias para ese mes
    filas = GuardiaCalendarioMensual.query.filter_by(
        hospital=hospital,
        anio=anio,
        mes=mes,
    ).all()

    asign_por_medico = {f.medico_id: f for f in filas}

    # Construimos filas por médico
    medicos_rows = []
    for m in medicos:
        asign = asign_por_medico.get(m.id)
        dias_list = asign.dias_list if asign else []

        dias_list_sorted = sorted(dias_list)
        total_guardias = len(dias_list_sorted)

        # Solo médicos con guardias asignadas
        if total_guardias == 0:
            continue

        dias_str = ", ".join(str(d) for d in dias_list_sorted)

        medicos_rows.append({
            "medico": m,
            "dias_str": dias_str,
            "total": total_guardias,
        })

    # Ordenar por el primer día de guardia
    def primer_dia(row):
        dias = [int(x.strip()) for x in row["dias_str"].split(",") if x.strip().isdigit()]
        return min(dias) if dias else 999

    medicos_rows.sort(key=primer_dia)

    # Etiqueta de mes
    meses_es = [
        "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO",
        "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE",
    ]
    mes_label = f"{meses_es[mes - 1]} {anio}"

    html = render_template(
        "guardias_calendario_pdf.html",
        hospital=hospital,
        hosp=hosp,
        mes_label=mes_label,
        medicos_rows=medicos_rows,
        generated_at=now_local(),
    )

    filename = f"guardias_calendario_{anio}{mes:02d}.pdf"
    response = render_pdf_from_html(html, pdf_filename=filename)
    return response


# ======================================================================
#   PWA: manifest / service worker / offline
# ======================================================================
@app.route("/manifest.webmanifest")
def manifest():
    return send_from_directory(
        "static", "manifest.webmanifest", mimetype="application/manifest+json"
    )


@app.route("/sw.js")
def sw():
    return send_from_directory("static", "sw.js", mimetype="application/javascript")


@app.route("/offline")
def offline():
    return render_template("offline.html")


# ==============================
# MAIN
# ==============================
if __name__ == "__main__":
    port = int(os.getenv("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)
